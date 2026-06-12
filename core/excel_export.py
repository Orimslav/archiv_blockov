"""Excel / CSV export (openpyxl + stdlib csv)."""

import csv
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from models.models import Profile, Receipt

_MONTHS_SK = [
    "", "Január", "Február", "Marec", "Apríl", "Máj", "Jún",
    "Júl", "August", "September", "Október", "November", "December",
]

_HEADERS_SIMPLE = ["Dátum", "Predajca", "Kategória", "Celkom", "Platba", "Popis"]
_HEADERS_VAT = [
    "Dátum", "Predajca", "Kategória", "0% Základ", "5% Z", "5% D",
    "19% Z", "19% D", "23% Z", "23% D", "Zaokr.", "Celkom", "Platba", "Popis",
]

# 0-based indices of monetary columns (kept as real numbers, not text).
_NUMERIC_COLS_SIMPLE = frozenset({3})
_NUMERIC_COLS_VAT = frozenset({3, 4, 5, 6, 7, 8, 9, 10, 11})

# Excel number format → Slovak locale renders it with a decimal comma.
_NUM_FMT = "#,##0.00"


def _numeric_cols(vat_enabled: bool) -> frozenset:
    return _NUMERIC_COLS_VAT if vat_enabled else _NUMERIC_COLS_SIMPLE


def _row_values(r: Receipt, vat_enabled: bool) -> list:
    """Row with monetary cells as real numbers (not formatted strings)."""
    platba = "Karta" if r.platba == "karta" else "Hotovosť"
    category = r.category_name or "Nezaradené"
    if not vat_enabled:
        return [
            r.datum_display(), r.vendor_name or "", category,
            round(r.celkom or 0, 2), platba, r.popis or "",
        ]
    return [
        r.datum_display(), r.vendor_name or "", category,
        round(r.base_0 or 0, 2), round(r.base_5 or 0, 2), round(r.tax_5 or 0, 2),
        round(r.base_19 or 0, 2), round(r.tax_19 or 0, 2),
        round(r.base_23 or 0, 2), round(r.tax_23 or 0, 2),
        round(r.zaokruhlenie or 0, 2), round(r.celkom or 0, 2), platba, r.popis or "",
    ]


def _sk_decimal(value: float) -> str:
    """Format a number with two decimals and a Slovak decimal comma."""
    return f"{value:.2f}".replace(".", ",")


def _column_sums(rows: List[list], numeric: frozenset) -> dict:
    """Sum each numeric column across all data rows (rounded to 2 decimals)."""
    return {i: round(sum(row[i] for row in rows), 2) for i in numeric}


def export_period_csv(path: Path, receipts: List[Receipt], profile: Profile) -> Path:
    """Write receipts to a UTF-8-BOM, ';'-separated CSV (Slovak locale)."""
    headers = _HEADERS_VAT if profile.vat_enabled else _HEADERS_SIMPLE
    numeric = _numeric_cols(profile.vat_enabled)
    rows = [_row_values(r, profile.vat_enabled) for r in receipts]
    sums = _column_sums(rows, numeric)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([
                _sk_decimal(v) if i in numeric else v for i, v in enumerate(row)
            ])
        # SPOLU row sums every numeric column (full VAT breakdown + Celkom).
        writer.writerow([])
        total_row = [""] * len(headers)
        total_row[0] = "SPOLU"
        for i, value in sums.items():
            total_row[i] = _sk_decimal(value)
        writer.writerow(total_row)
    return path


def export_period_xlsx(path: Path, receipts: List[Receipt], profile: Profile) -> Path:
    """Write receipts to an .xlsx workbook with a totals row.

    Monetary values are stored as real numbers with a Slovak number format
    (decimal comma) so they stay summable in Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    headers = _HEADERS_VAT if profile.vat_enabled else _HEADERS_SIMPLE
    numeric = _numeric_cols(profile.vat_enabled)
    rows = [_row_values(r, profile.vat_enabled) for r in receipts]
    wb = Workbook()
    ws = wb.active
    ws.title = "Bločky"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    # Apply the Slovak number format to the monetary columns of the data rows.
    for row_cells in ws.iter_rows(min_row=2, max_row=1 + len(rows)):
        for idx in numeric:
            row_cells[idx].number_format = _NUM_FMT

    # SPOLU row sums every numeric column (full VAT breakdown + Celkom).
    sums = _column_sums(rows, numeric)
    ws.append([])
    total_row = [""] * len(headers)
    total_row[0] = "SPOLU"
    for i, value in sums.items():
        total_row[i] = value
    ws.append(total_row)
    last = ws.max_row
    ws.cell(row=last, column=1).font = Font(bold=True)
    for idx in numeric:
        cell = ws.cell(row=last, column=idx + 1)
        cell.font = Font(bold=True)
        cell.number_format = _NUM_FMT
    wb.save(path)
    return path


def export_item_report_xlsx(
    path: Path,
    item_label: str,
    monthly: Sequence[Tuple[int, float, float]],
    profile: Profile,
    year: Optional[int] = None,
) -> Path:
    """Write a per-month item consumption report (month, quantity, spend)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Spotreba"

    scope = f"Rok {year}" if year else "Všetky obdobia"
    title = ws.cell(row=1, column=1, value=f"Spotreba: {item_label}")
    title.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Profil: {profile.name}   ·   {scope}")

    header_row = 4
    headers = ["Mesiac", "Množstvo", "Suma (€)"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = Font(bold=True)

    total_qty = total_spend = 0.0
    row = header_row + 1
    for month, qty, spend in monthly:
        name = _MONTHS_SK[month] if 1 <= month <= 12 else "—"
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(qty, 3)).number_format = "#,##0.###"
        ws.cell(row=row, column=3, value=round(spend, 2)).number_format = _NUM_FMT
        total_qty += qty
        total_spend += spend
        row += 1

    ws.cell(row=row, column=1, value="SPOLU").font = Font(bold=True)
    qty_cell = ws.cell(row=row, column=2, value=round(total_qty, 3))
    qty_cell.font = Font(bold=True)
    qty_cell.number_format = "#,##0.###"
    spend_cell = ws.cell(row=row, column=3, value=round(total_spend, 2))
    spend_cell.font = Font(bold=True)
    spend_cell.number_format = _NUM_FMT

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for r in ws.iter_rows(min_row=header_row, min_col=2, max_col=3):
        for cell in r:
            cell.alignment = Alignment(horizontal="right")

    wb.save(path)
    return path
